from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


driver = webdriver.Chrome("./chromedriver.exe")
wait = WebDriverWait(driver=driver, timeout=60)

#読み込みexcelファイル
wb = openpyxl.load_workbook("ip.xlsx")
ws = wb.worksheets[0]
#書き込みexcelファイル
wb2 = openpyxl.load_workbook("result.xlsx")
ws2 = wb2.worksheets[0]

ip_list = []
index = 2

for col in ws["A"]:
  ip_list.append(col.value.splitlines()[0])

blacklistalert_url = "https://blacklistalert.org/"
blacklistalert_form_xpath = '/html/body/center/font/form/input[1]'
click_button_xpath = '/html/body/center/font/form/input[3]'

spamcop_url = "https://www.spamcop.net/bl.shtml"
spamcop_result_xpath = "//*[@id='content']/p"
spamcop_form_selector = "#content > form input[name='ip']"
spamcop_click_selector = "#content > form input[type=submit]"

trendmicro_url = 'https://servicecentral.trendmicro.com/en-US/ers/ip-lookup/'
trendmicro_result_selector = '.listedin.detail'
trendmicro_form_selector = '#ip-form'
trendmicro_click_selector = '#check-ip'

for ip in ip_list:
  ##各サイト巡回
  driver.get(blacklistalert_url)
  driver.implicitly_wait(1)
  form = driver.find_element(By.XPATH, blacklistalert_form_xpath)
  form.clear()
  form.send_keys(ip)
  click_button = driver.find_element(By.XPATH,click_button_xpath)
  click_button.click()
  wait.until(EC.text_to_be_present_in_element((By.TAG_NAME, 'body'),'sponsored'))
  ip_tables = driver.find_elements(By.TAG_NAME,'table')
  trs = ip_tables[0].find_elements(By.TAG_NAME, "tr")
  
  NG_count = 0
  NG_sites = []
  trs_count = 0
  for tr in trs:
    trs_count += 1
    site = tr.find_element(By.CLASS_NAME,'left').text
    status = tr.find_elements(By.TAG_NAME,"strong")
    if len(status) == 0:
      status = "NG"
    else:
      status = "OK"
    if status == "NG":
      NG_count += 1
      NG_sites.append(site)
  #table3つ目が存在する場合これも追加
  if len(ip_tables) == 3:
    trs = ip_tables[2].find_elements(By.TAG_NAME, "tr")
    for tr in trs:
      trs_count += 1
      site = tr.find_element(By.CLASS_NAME,'left').text
      status = tr.find_elements(By.TAG_NAME,"strong")
      if len(status) == 0:
        status = "NG"
      else:
        status = "OK"
      if status == "NG":
        NG_count += 1
        NG_sites.append(site)
  print(f'ip:{ip}, blacklistalert_count:{trs_count}')
  #spamcop
  driver.get(spamcop_url)
  driver.implicitly_wait(1)
  spamcop_form = driver.find_element(By.CSS_SELECTOR, spamcop_form_selector)
  spamcop_form.send_keys(ip)
  spamcop_click_button = driver.find_element(By.CSS_SELECTOR,spamcop_click_selector)
  spamcop_click_button.click()
  spamcop_result = driver.find_elements(By.XPATH,spamcop_result_xpath)[0].text
  #trendmicro
  driver.get(trendmicro_url)
  driver.implicitly_wait(1)
  trendmicro_form = driver.find_element(By.CSS_SELECTOR, trendmicro_form_selector)
  trendmicro_form.send_keys(ip)
  trendmicro_click_button = driver.find_element(By.CSS_SELECTOR,trendmicro_click_selector)
  trendmicro_click_button.click()
  wait.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR,trendmicro_result_selector),'List'))
  trendmicro_result = driver.find_element(By.CSS_SELECTOR,trendmicro_result_selector).text

  ##Excel書き込み
  ws2.cell(row=index,column=1).value = ip
  #blacklistalert
  if NG_count != 0:
    ws2.cell(row=index,column=3).value = ", ".join(NG_sites)
    ws2.cell(row=index,column=2).value = "NG"
    ws2[f'B{index}'].font = openpyxl.styles.fonts.Font(color='FF0000')
  else:
    ws2.cell(row=index,column=2).value = "OK"
    ws2[f'B{index}'].font = openpyxl.styles.fonts.Font(color='00793D')
  #spamcop
  if 'not listed in' in spamcop_result:
    ws2.cell(row=index,column=4).value = "OK"
    ws2[f'D{index}'].font = openpyxl.styles.fonts.Font(color='00793D')
  else:
    ws2.cell(row=index,column=4).value = "NG"
    ws2[f'D{index}'].font = openpyxl.styles.fonts.Font(color='FF0000')
  #trendmicro
  if 'None' in trendmicro_result:
    ws2.cell(row=index,column=5).value = "OK"
    ws2[f'E{index}'].font = openpyxl.styles.fonts.Font(color='00793D')
  else:
    ws2.cell(row=index,column=5).value = "NG"
    ws2[f'E{index}'].font = openpyxl.styles.fonts.Font(color='FF0000')
    ws2.cell(row=index,column=6).value = trendmicro_result
  wb2.save("./result.xlsx")
  index += 1
  
